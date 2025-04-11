from django.http import HttpResponse
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from django.views.decorators.csrf import csrf_exempt
from io import BytesIO
from openpyxl.styles import Font
         
@csrf_exempt
def FormatoArchivo(request):    
   try:
      Archivo = request.FILES['archivo']
         
      if Archivo.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":

         # Abrir el archivo Excel
         wb = load_workbook(Archivo)
         ws = wb.active

         # Añadir el logo
         logo = Image('moduloetl/Unicaribe.png') 
         logo.height = 150
         logo.width = 400
         ws.merge_cells('A1:C8')
         ws.add_image(logo, 'A1') 

         # Añadir encabezado al archivo excel
         ws['I2'] = "UNIVERSIDAD DEL CARIBE"
         ws['I3'] = "SECRETARÍA DE PLANEACIÓN Y DESARROLLO INSTITUCIONAL"
         ws['I4'] = "DEPARTAMENTO DE CONTROL Y EVALUACIÓN"
         ws['I5'] = "INFORME DE ACTIVIDADES"
         ws['I6'] = Archivo.name.replace('.xlsx', '')

         # Ajustar el tamaño de fuente y el estilo
         ws['I2'].font = Font(size=12, bold=True)
         ws['I3'].font = Font(size=12, bold=True)
         ws['I4'].font = Font(size=12, bold=True)
         ws['I5'].font = Font(size=12, bold=True)
         ws['I6'].font = Font(size=12, bold=True)

         # Ajustar la alineación del texto
         ws['I2'].alignment = ws['I2'].alignment.copy(horizontal='center')
         ws['I3'].alignment = ws['I3'].alignment.copy(horizontal='center')
         ws['I4'].alignment = ws['I4'].alignment.copy(horizontal='center')
         ws['I5'].alignment = ws['I5'].alignment.copy(horizontal='center')
         ws['I6'].alignment = ws['I6'].alignment.copy(horizontal='center')

         # Crear un buffer para guardar el archivo
         buffer = BytesIO()
         # Guardar el archivo en el buffer
         wb.save(buffer)
         buffer.seek(0)
        
         # Devolver el archivo como respuesta
         response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
         )
         buffer.close()
         wb.close()
         return response
         
      else:
         return HttpResponse(
            json.dumps({'status': 'error', 'message': 'Tipo de archivo no soportado'}),
            content_type='application/json',
            status=400
         )
      
   except Exception as e:
     print("Error al procesar el archivo:", str(e))
     return HttpResponse(
        json.dumps({'status': 'error', 'message': str(e)}),
        content_type='application/json',
        status=400
      ) 
